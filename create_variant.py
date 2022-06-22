from google.cloud import dialogflowcx_v3
import asyncio
from openpyxl import Workbook
from openpyxl import load_workbook
import time

# export GOOGLE_APPLICATION_CREDENTIALS="heartschat-prod-a505-9599eda00cef.json"

ROUTE_MAP = "route_map_06-22-2022_00-25-45.xlsx"
ENDPOINT = "us-central1-dialogflow.googleapis.com"

async def create_variant():
    client = dialogflowcx_v3.AgentsAsyncClient(client_options={"api_endpoint": ENDPOINT})
    agent = dialogflowcx_v3.Agent()
    agent.display_name = "test2"
    agent.default_language_code = "en"
    agent.time_zone = "GMT-8:00"
    request = dialogflowcx_v3.CreateAgentRequest(parent="projects/heartschat-prod-a505/locations/us-central1", agent=agent)
    agent = await client.create_agent(request=request)
    return agent.name

async def export_reference():
    client = dialogflowcx_v3.AgentsAsyncClient(client_options={"api_endpoint": ENDPOINT})
    workbook = load_workbook(ROUTE_MAP)
    ref_sheet = workbook["Reference"]
    request = dialogflowcx_v3.ExportAgentRequest(name=str(ref_sheet["A1"].value))
    operation = await client.export_agent(request=request)
    response = await operation.result()
    return response.agent_content

async def restore_reference(agent_name, agent_content):
    client = dialogflowcx_v3.AgentsAsyncClient(client_options={"api_endpoint": ENDPOINT})
    request = dialogflowcx_v3.RestoreAgentRequest(name=agent_name, agent_content=agent_content)
    operation = await client.restore_agent(request=request)
    response = await operation.result()

async def get_route_groups(agent_name):
    client = dialogflowcx_v3.TransitionRouteGroupsAsyncClient(client_options={"api_endpoint": ENDPOINT})
    request = dialogflowcx_v3.ListTransitionRouteGroupsRequest(parent="{}/flows/00000000-0000-0000-0000-000000000000".format(agent_name))
    route_groups = await client.list_transition_route_groups(request=request)
    return route_groups

async def update_variant(agent_name, route_groups):
    workbook = load_workbook(ROUTE_MAP)
    async for route_group in route_groups:
        for route in route_group.transition_routes:
            for i in range(1, len(workbook.sheetnames) - 3):
                route_sheet = workbook[str(i)]
                if route.name == route_sheet["A1"].value:
                    for message in route.trigger_fulfillment.messages:
                        counter = 2
                        for j in range(0, len(message.text.text)):
                            if route_sheet["B{}".format(counter)].value is not None and route_sheet["B{}".format(counter)].value != "":
                                message.text.text[j] = str(route_sheet["B{}".format(counter)].value)
                                counter += 1
                            else: 
                                message.text.text.pop(j)
                                j -= 1
                        for j in range(counter, 999):
                            if route_sheet["B{}".format(j)].value is not None and route_sheet["B{}".format(j)].value != "":
                                message.text.text.append(str(route_sheet["B{}".format(j)].value))
                            else:
                                break
        print("hi")
        client = dialogflowcx_v3.TransitionRouteGroupsAsyncClient(client_options={"api_endpoint": ENDPOINT})
        request = dialogflowcx_v3.UpdateTransitionRouteGroupRequest(transition_route_group=route_group)
        response = await client.update_transition_route_group(request=request)
        time.sleep(1)

agent_name = asyncio.run(create_variant())
agent_content = asyncio.run(export_reference())
asyncio.run(restore_reference(agent_name, agent_content))
route_groups = asyncio.run(get_route_groups(agent_name))
asyncio.run(update_variant(agent_name, route_groups))